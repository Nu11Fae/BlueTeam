from __future__ import annotations

import json
import re
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .hld_baseline import EXECUTIVE_HEADERS, RISK_REGISTER_HEADERS
from .models import FINDING_ID_PATTERN, Finding
from .scoring import assign_grade, classify_finding, final_score, inherent_risk, residual_risk, transaction_materiality
from .settings import AppSettings, get_settings


FIELD_BY_HEADER = {
    "Finding ID": "finding_id",
    "Title": "title",
    "Finding Type": "finding_type",
    "Affected Asset": "affected_asset",
    "Taxonomy": "taxonomy_display",
    "Description": "description",
    "Evidence Reference / Evidence Summary": "evidence_summary",
    "Validation Status": "validation_status",
    "Evidence Confidence": "evidence_confidence",
    "Likelihood": "likelihood",
    "Technical Impact": "technical_impact",
    "Business Impact": "business_impact",
    "Control Weakness": "control_weakness",
    "Compliance Exposure": "compliance_exposure",
    "Remediation Effort": "remediation_effort",
    "Transaction Impact": "transaction_impact",
    "Inherent Risk": "inherent_risk",
    "Residual Risk": "residual_risk",
    "Transaction Materiality": "transaction_materiality",
    "Final Score": "final_score",
    "Grade": "grade",
    "Classification (Red Flag / Negotiation Relevant / Integration Item / Observation)": "classification",
    "Suggested Action": "suggested_action",
}


def _normalize_text(value: object, fallback: str = "") -> str:
    text = str(value or fallback).strip()
    return re.sub(r"\s+", " ", text)


def _extract_json_payload(raw_text: str) -> dict[str, object]:
    text = raw_text.strip()
    if not text:
        return {}
    try:
        payload = json.loads(text)
        if isinstance(payload, dict):
            return payload
    except json.JSONDecodeError:
        pass
    fenced = re.search(r"```(?:json)?\s*(\{.*\})\s*```", text, flags=re.DOTALL)
    if fenced:
        try:
            payload = json.loads(fenced.group(1))
            if isinstance(payload, dict):
                return payload
        except json.JSONDecodeError:
            return {}
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end > start:
        try:
            payload = json.loads(text[start : end + 1])
            if isinstance(payload, dict):
                return payload
        except json.JSONDecodeError:
            return {}
    return {}


def _tool_match_candidates(tool_findings: dict[str, list[dict[str, object]]], affected_asset: str) -> list[dict[str, object]]:
    normalized_asset = affected_asset.strip().lower()
    if not normalized_asset:
        return []
    matches: list[dict[str, object]] = []
    for asset_path, entries in tool_findings.items():
        candidate = asset_path.lower()
        if candidate == normalized_asset or candidate.endswith(normalized_asset) or normalized_asset.endswith(candidate):
            matches.extend(entries)
    return matches


def _finding_from_raw(raw: dict[str, object], index: int) -> Finding:
    if not FINDING_ID_PATTERN.fullmatch(_normalize_text(raw.get("finding_id") or raw.get("id"), "").upper()):
        raw = dict(raw)
        raw["finding_id"] = f"TLF-DA-{index:03d}"
    return Finding.model_validate(raw)


def _apply_tool_validation(finding: Finding, tool_findings: dict[str, list[dict[str, object]]]) -> Finding:
    matches = _tool_match_candidates(tool_findings, finding.affected_asset)
    if matches:
        sources = []
        for match in matches:
            tool_name = _normalize_text(match.get("tool"), "tool")
            asset = _normalize_text(match.get("path"), finding.affected_asset)
            title = _normalize_text(match.get("title") or match.get("rule") or match.get("summary"), "anchored evidence")
            sources.append(f"{tool_name}: {asset} :: {title}")
        finding.evidence_sources = sorted(set(sources))
        finding.validation_status = "Validated"
        finding.llm_only = False
        return finding
    finding.validation_status = "Candidate"
    finding.llm_only = True
    return finding


def _apply_validation_feedback(findings: list[Finding], payload: dict[str, object]) -> list[Finding]:
    by_id = {finding.finding_id: finding for finding in findings}
    drops = {
        _normalize_text(item.get("drop") or item.get("finding_id")).upper()
        for item in payload.get("duplicates", [])
        if isinstance(item, dict)
    }
    for item in payload.get("inconsistencies", []):
        if not isinstance(item, dict):
            continue
        finding = by_id.get(_normalize_text(item.get("finding_id")).upper())
        if not finding:
            continue
        updates = item.get("suggested_updates", {})
        if not isinstance(updates, dict):
            continue
        for field_name in (
            "validation_status",
            "evidence_confidence",
            "suggested_action",
            "affected_asset",
            "taxonomy",
            "likelihood",
            "technical_impact",
            "business_impact",
            "control_weakness",
            "compliance_exposure",
            "remediation_effort",
            "transaction_impact",
        ):
            if field_name in updates:
                setattr(finding, field_name, updates[field_name])
        by_id[finding.finding_id] = Finding.model_validate(finding.model_dump())
    for item in payload.get("missing_taxonomy", []):
        if not isinstance(item, dict):
            continue
        finding = by_id.get(_normalize_text(item.get("finding_id")).upper())
        if not finding:
            continue
        if not finding.taxonomy:
            suggested = item.get("suggested_taxonomy") or item.get("taxonomy") or []
            hydrated = Finding.model_validate(
                {
                    "finding_id": finding.finding_id,
                    "title": finding.title,
                    "finding_type": finding.finding_type,
                    "affected_asset": finding.affected_asset,
                    "taxonomy": suggested,
                    "description": finding.description,
                    "evidence_summary": finding.evidence_summary,
                }
            )
            finding.taxonomy = hydrated.taxonomy
        if not finding.taxonomy:
            finding.validation_status = "Candidate"
    return [item for item in by_id.values() if item.finding_id not in drops]


def _extract_raw_findings(path: Path) -> list[dict[str, object]]:
    payload = _extract_json_payload(path.read_text(encoding="utf-8")) if path.exists() else {}
    for key in ("validated_findings", "deep_review_findings", "findings"):
        raw_findings = payload.get(key)
        if isinstance(raw_findings, list):
            return [item for item in raw_findings if isinstance(item, dict)]
    return []


def _dedupe_key(finding: Finding) -> tuple[object, ...]:
    return (
        finding.title.lower(),
        finding.finding_type,
        finding.affected_asset.lower(),
        tuple(sorted(item.lower() for item in finding.taxonomy)),
    )


def _finding_rank(finding: Finding, settings: AppSettings) -> tuple[object, ...]:
    confidence_rank = {"Low": 0, "Medium": 1, "High": 2}
    return (
        finding.validation_status == "Validated",
        confidence_rank.get(finding.evidence_confidence, 0),
        len(finding.evidence_sources),
        len(finding.taxonomy),
        final_score(finding, settings),
    )


def _deduplicate_findings(findings: list[Finding], settings: AppSettings) -> list[Finding]:
    by_key: dict[tuple[object, ...], Finding] = {}
    for finding in findings:
        key = _dedupe_key(finding)
        current = by_key.get(key)
        if current is None or _finding_rank(finding, settings) > _finding_rank(current, settings):
            by_key[key] = finding
    return list(by_key.values())


def _finding_record(finding: Finding, settings: AppSettings) -> dict[str, object]:
    score = final_score(finding, settings)
    grade = assign_grade(score, settings)
    classification = classify_finding(finding, settings)
    return {
        **finding.model_dump(),
        "taxonomy_display": ", ".join(finding.taxonomy),
        "inherent_risk": inherent_risk(finding, settings),
        "residual_risk": residual_risk(finding, settings),
        "transaction_materiality": transaction_materiality(finding, settings),
        "final_score": score,
        "grade": grade,
        "classification": classification,
    }


def load_llm_findings(
    path: Path,
    *,
    tool_findings: dict[str, list[dict[str, object]]] | None = None,
    validation_path: Path | None = None,
    settings: AppSettings | None = None,
    supplemental_paths: list[Path] | None = None,
) -> list[Finding]:
    resolved_settings = settings or get_settings()
    raw_findings: list[dict[str, object]] = []
    raw_findings.extend(_extract_raw_findings(path))
    for supplemental_path in supplemental_paths or []:
        raw_findings.extend(_extract_raw_findings(supplemental_path))
    findings = [
        _apply_tool_validation(_finding_from_raw(raw, index), tool_findings or {})
        for index, raw in enumerate(raw_findings, start=1)
    ]
    findings = _deduplicate_findings(findings, resolved_settings)
    if validation_path and validation_path.exists():
        findings = _apply_validation_feedback(findings, _extract_json_payload(validation_path.read_text(encoding="utf-8")))
    findings = [Finding.model_validate(item.model_dump()) for item in findings]
    findings.sort(
        key=lambda item: (final_score(item, resolved_settings), item.validation_status == "Validated", item.evidence_confidence == "High"),
        reverse=True,
    )
    return findings


def _style_row_map(worksheet, row_index: int, max_columns: int) -> dict[int, object]:
    return {index: copy(worksheet.cell(row=row_index, column=index)._style) for index in range(1, max_columns + 1)}


def _apply_row_style(worksheet, row_index: int, styles: dict[int, object], max_columns: int) -> None:
    for index in range(1, max_columns + 1):
        cell = worksheet.cell(row=row_index, column=index)
        cell._style = copy(styles[index])
        cell.alignment = copy(cell.alignment) if cell.alignment else Alignment(wrap_text=True)


def _clear_row(worksheet, row_index: int, max_columns: int) -> None:
    for index in range(1, max_columns + 1):
        worksheet.cell(row=row_index, column=index).value = None


def _is_vulnerability(record: dict[str, object]) -> bool:
    finding_type = str(record.get("finding_type", "")).lower()
    taxonomy = str(record.get("taxonomy_display", "")).lower()
    return "vulnerability" in finding_type or any(token in taxonomy for token in ("cwe", "cve", "owasp", "capec"))


def build_risk_register_workbook(template_path: Path, output_path: Path, findings: list[Finding], settings: AppSettings | None = None) -> Path:
    resolved_settings = settings or get_settings()
    workbook = load_workbook(template_path)
    register = workbook["Risk Register"]
    executive = workbook["Executive Vulnerabilities"]
    scoring_sheet = workbook["Scoring Parameters"] if "Scoring Parameters" in workbook.sheetnames else None

    register_styles = _style_row_map(register, 10, len(RISK_REGISTER_HEADERS))
    executive_styles = _style_row_map(executive, 7, len(EXECUTIVE_HEADERS))
    _clear_row(register, 10, len(RISK_REGISTER_HEADERS))
    _clear_row(executive, 7, len(EXECUTIVE_HEADERS))

    register_headers = [cell.value for cell in register[9][: len(RISK_REGISTER_HEADERS)]]
    normalized_records = [_finding_record(finding, resolved_settings) for finding in findings]
    for offset, record in enumerate(normalized_records, start=10):
        if offset > 10:
            register.insert_rows(offset)
        _apply_row_style(register, offset, register_styles, len(RISK_REGISTER_HEADERS))
        for column_index, header in enumerate(register_headers, start=1):
            field_name = FIELD_BY_HEADER.get(str(header))
            if field_name:
                register.cell(row=offset, column=column_index, value=record.get(field_name))

    executive_rows = [record for record in normalized_records if _is_vulnerability(record)]
    executive_headers = [cell.value for cell in executive[6][: len(EXECUTIVE_HEADERS)]]
    for offset, record in enumerate(executive_rows, start=7):
        if offset > 7:
            executive.insert_rows(offset)
        _apply_row_style(executive, offset, executive_styles, len(EXECUTIVE_HEADERS))
        executive_mapping = {
            "Finding ID": record.get("finding_id"),
            "Title": record.get("title"),
            "Affected Asset": record.get("affected_asset"),
            "Taxonomy / Standard Mapping": record.get("taxonomy_display"),
            "Validation Status": record.get("validation_status"),
            "Evidence Confidence": record.get("evidence_confidence"),
            "Residual Risk": record.get("residual_risk"),
            "Final Score": record.get("final_score"),
            "Grade": record.get("grade"),
            "Classification": record.get("classification"),
            "Executive Description": record.get("executive_description"),
            "Key Evidence": record.get("key_evidence"),
            "Suggested Action": record.get("suggested_action"),
        }
        for column_index, header in enumerate(executive_headers, start=1):
            if header:
                executive.cell(row=offset, column=column_index, value=executive_mapping.get(str(header)))

    if scoring_sheet:
        scoring_sheet["A2"] = "Settings Hash"
        scoring_sheet["B2"] = resolved_settings.settings_hash()
        row = 4
        for section, weights in resolved_settings.scoring_weights.items():
            scoring_sheet.cell(row=row, column=1, value=section)
            row += 1
            for key, value in weights.items():
                scoring_sheet.cell(row=row, column=1, value=key)
                scoring_sheet.cell(row=row, column=2, value=value)
                row += 1
        row += 1
        scoring_sheet.cell(row=row, column=1, value="Grade Thresholds")
        row += 1
        for item in resolved_settings.grade_thresholds:
            scoring_sheet.cell(row=row, column=1, value=item["grade"])
            scoring_sheet.cell(row=row, column=2, value=item["min"])
            scoring_sheet.cell(row=row, column=3, value=item["max"])
            row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def export_risk_register(
    template_path: Path,
    output_dir: Path,
    findings: list[Finding],
    settings: AppSettings | None = None,
) -> tuple[Path, Path]:
    resolved_settings = settings or get_settings()
    workbook_path = build_risk_register_workbook(template_path, output_dir / "Risk Register.xlsx", findings, resolved_settings)
    normalized_records = [_finding_record(finding, resolved_settings) for finding in findings]
    payload = {
        "risk_register": normalized_records,
        "scoring_parameters": {
            "weights": resolved_settings.scoring_weights,
            "grade_thresholds": resolved_settings.grade_thresholds,
            "classification_rules": resolved_settings.classification_rules,
            "settings_hash": resolved_settings.settings_hash(),
        },
        "executive_vulnerabilities": [record for record in normalized_records if _is_vulnerability(record)],
    }
    json_path = output_dir / "risk_register.json"
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    return workbook_path, json_path


def build_risk_score_artifact(output_path: Path, findings: list[Finding], settings: AppSettings | None = None) -> Path:
    resolved_settings = settings or get_settings()
    records = [_finding_record(finding, resolved_settings) for finding in findings]
    grade_histogram = Counter(str(item.get("grade", "N/A")) for item in records)
    validation_histogram = Counter(str(item.get("validation_status", "Candidate")) for item in records)
    dimension_keys = [
        "likelihood",
        "technical_impact",
        "business_impact",
        "control_weakness",
        "compliance_exposure",
        "remediation_effort",
        "transaction_impact",
    ]
    lines = [
        "# Risk Score Artifact",
        "",
        "## Scoring Model",
        "",
        "- Inherent Risk = 0.35 x Likelihood + 0.40 x Technical Impact + 0.25 x Business Impact",
        "- Residual Risk = 0.70 x Inherent Risk + 0.30 x Control Weakness",
        "- Transaction Materiality = 0.50 x Transaction Impact + 0.25 x Compliance Exposure + 0.25 x Remediation Effort",
        "- Final Score = max(Residual Risk, Transaction Materiality)",
        "- Grade bands: " + ", ".join(
            f"{item['grade']} {float(item['min']):.1f}-{float(item['max']):.1f}"
            for item in resolved_settings.grade_thresholds
        ),
        "",
        f"- Findings exported: {len(records)}",
        f"- Validated findings: {validation_histogram.get('Validated', 0)}",
        f"- Candidate findings: {validation_histogram.get('Candidate', 0)}",
        "",
        "## Grade Distribution",
        "",
        "| Grade | Count |",
        "| --- | ---: |",
    ]
    for threshold in resolved_settings.grade_thresholds:
        grade = str(threshold["grade"])
        lines.append(f"| {grade} | {grade_histogram.get(grade, 0)} |")
    lines.extend(["", "## Eight-Dimension Overview", "", "| Dimension | Average | Max |", "| --- | ---: | ---: |"])
    for key in dimension_keys:
        values = [int(item[key]) for item in records if key in item]
        average = (sum(values) / len(values)) if values else 0.0
        maximum = max(values) if values else 0
        lines.append(f"| {key.replace('_', ' ').title()} | {average:.2f} | {maximum} |")
    lines.extend(["", "## Top Findings", "", "| Finding ID | Title | Final Score | Grade | Classification | Validation |", "| --- | --- | ---: | --- | --- | --- |"])
    for record in records[:20]:
        lines.append(
            f"| {record['finding_id']} | {record['title']} | {record['final_score']:.2f} | {record['grade']} | "
            f"{record['classification']} | {record['validation_status']} |"
        )
    output_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")
    return output_path
