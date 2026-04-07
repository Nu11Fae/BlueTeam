from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


def _looks_like_date_header(header: str | None) -> bool:
    if not header:
        return False
    lowered = header.lower()
    return any(token in lowered for token in ("start", "end", "date", "window"))


def _normalize_cell(value: object, header: str | None = None) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and _looks_like_date_header(header):
        try:
            converted = from_excel(value)
        except (TypeError, ValueError, OverflowError):
            return value
        if isinstance(converted, datetime):
            if converted.hour == 0 and converted.minute == 0 and converted.second == 0 and converted.microsecond == 0:
                return converted.date().isoformat()
            return converted.isoformat()
        if isinstance(converted, date):
            return converted.isoformat()
    return value


def parse_risk_register(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=True)
    register = workbook["Risk Register"] if "Risk Register" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    executive = workbook["Executive Vulnerabilities"] if "Executive Vulnerabilities" in workbook.sheetnames else None

    headers = [cell.value for cell in register[9]]
    rows = []
    for row in register.iter_rows(min_row=10, values_only=True):
        if not any(row):
            continue
        record = {
            str(headers[index]): _normalize_cell(value, str(headers[index]))
            for index, value in enumerate(row)
            if headers[index]
        }
        rows.append(record)

    executive_rows: list[dict[str, object]] = []
    if executive:
        executive_headers = [cell.value for cell in executive[6]]
        for row in executive.iter_rows(min_row=7, values_only=True):
            if not any(row):
                continue
            executive_rows.append(
                {
                    str(executive_headers[index]): _normalize_cell(value, str(executive_headers[index]))
                    for index, value in enumerate(row)
                    if executive_headers[index]
                }
            )

    return {
        "sheet_names": workbook.sheetnames,
        "entries": rows,
        "executive_entries": executive_rows,
        "total_entries": len(rows),
        "critical_entries": sum(1 for row in rows if row.get("Grade") == "E"),
    }


def parse_planning_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headline = _normalize_cell(sheet["A1"].value, "headline")
    window = _normalize_cell(sheet["A2"].value, "window")
    author_line = _normalize_cell(sheet["A3"].value, "date")
    header_row = None
    header_index = None
    for idx in range(1, min(sheet.max_row, 12) + 1):
        candidate = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[idx]]
        lowered = {value.lower() for value in candidate if value}
        if {"task", "start", "end"} & lowered or ("planned start" in lowered and "planned end" in lowered):
            header_row = candidate
            header_index = idx
            break
    if header_row is None and "Gantt" in workbook.sheetnames:
        sheet = workbook["Gantt"]
        header_index = 9
        header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[header_index]]
    tasks: list[dict[str, object]] = []
    if header_row and header_index:
        for row in sheet.iter_rows(min_row=header_index + 1, values_only=True):
            if not any(row):
                continue
            record = {
                header_row[index]: _normalize_cell(value, header_row[index])
                for index, value in enumerate(row)
                if index < len(header_row) and header_row[index]
            }
            if not record.get("WBS") and not record.get("Task") and not record.get("Planned Start"):
                continue
            tasks.append(record)
    return {
        "sheet_name": sheet.title,
        "headline": headline,
        "window": window,
        "author_line": author_line,
        "tasks": tasks,
        "task_count": len(tasks),
    }
